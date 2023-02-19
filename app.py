from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
import xml.etree.ElementTree as ET
import locale
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/', methods=['POST'])
def convert():
    xml_files = request.files.getlist("xml_files")
    if not xml_files:
        return "No files selected"
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    wb = Workbook()
    ws = wb.active
    row = 1
    # Add the column headers
    ws.cell(row=row, column=1).value = "Product Code"
    ws.cell(row=row, column=2).value = "Product Name"
    ws.cell(row=row, column=3).value = "Product Value"
    ws.cell(row=row, column=4).value = "Product Quantity"
    ws.cell(row=row, column=5).value = "Product Unit Value"
    ws.cell(row=row, column=6).value = "Product Total Value"
    row += 1

    for xml_file in xml_files:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        for det in root.iter("det"):
            for child in det.iter("prod"):
                ws.cell(row=row, column=1).value = child.find("cProd").text
                ws.cell(row=row, column=2).value = child.find("xProd").text
                ws.cell(row=row, column=3).value = locale.currency(float(child.find("vProd").text), grouping=True)
                ws.cell(row=row, column=4).value = f"{float(child.find('qCom').text):.2f}"
                ws.cell(row=row, column=5).value = locale.currency(float(child.find("vUnCom").text), grouping=True)
                ws.cell(row=row, column=6).value = locale.currency(float(child.find("vProd").text), grouping=True)
                row +=1

    xlsx_file = "relatorio.xlsx"
    wb.save(xlsx_file)
    return send_file(xlsx_file, as_attachment=True)

if __name__ == '__main__':
    app.run()
